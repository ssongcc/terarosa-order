"""
테라로사 자사몰 주문취합 처리 스크립트
사용법: python terrarosa_order_processor.py <주문취합엑셀.xlsx> <자사몰상품코드.xlsx>
"""

import re
import sys
from datetime import datetime
from copy import copy

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────
REMOVE_STRINGS = [
    "/구매 안함", "/플러스", "[플러스] ", "/불필요", "/필요",
    "불필요", "필요", "/상자 없음", "/포장 없음",
    "테라로사 시그니처 ", "[Online Exclusive] ", "[Online Exclusive/플러스] ",
    "[C.O.E/플러스] ",   #
]

TEXT_REPLACE = {
    "중간 분쇄(드립용)": "드립용",
    "가는 분쇄(에스프레소용)": "에스프레소용",
}

# 색상
COLOR_DRIP = "E2EFDA"
COLOR_SCOOP = "FFF2CC"
COLOR_HEADER = "D9D9D9"
COLOR_WHITE = "FFFFFF"

# 열 너비
COL_WIDTHS = {"A": 42, "B": 10, "C": 35, "D": 8, "E": 18}

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ──────────────────────────────────────────────
# 1. 데이터 로드
# ──────────────────────────────────────────────
def load_order_data(filepath: str) -> pd.DataFrame:
    """'취합용' 시트에서 A(품목명), B(수량) 읽기"""
    df = pd.read_excel(filepath, sheet_name="취합용", header=0, dtype=str)
    df = df.iloc[:, :2].copy()
    df.columns = ["품목명_원본", "수량"]
    df.dropna(subset=["품목명_원본"], inplace=True)
    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    return df.reset_index(drop=True)


def load_code_data(filepath: str) -> pd.DataFrame:
    """자사몰상품코드 파일 로드 (첫 번째 시트)"""
    df = pd.read_excel(filepath, header=0, dtype=str)
    df.fillna("", inplace=True)
    return df


# ──────────────────────────────────────────────
# 2. A열 텍스트 정리
# ──────────────────────────────────────────────
def clean_item_name(name: str) -> str:
    """제거 문자열 처리 및 텍스트 변경"""
    for s in REMOVE_STRINGS:
        name = name.replace(s, "")
    for old, new in TEXT_REPLACE.items():
        name = name.replace(old, new)
    # 어센틱 정기 배송 → 어센틱 에스프레소 블렌드
    if "어센틱" in name and "정기 배송" in name:
        # 언더스코어 앞 부분만 교체 (옵션은 유지)
        if "_" in name:
            option = name.split("_", 1)[1]
            name = "어센틱 에스프레소 블렌드_" + option
        else:
            name = "어센틱 에스프레소 블렌드"
    return name.strip()


def apply_sos_weight(name: str, weight: str) -> str:
    """S.O.S 포함 품목 중량 300g→150g"""
    if "S.O.S" in name and weight == "300g":
        return "150g"
    return weight


# ──────────────────────────────────────────────
# 3. A열 분리 (품목명 / 옵션 / 중량 추출)
# ──────────────────────────────────────────────
WEIGHT_PATTERN = re.compile(r"(\d+(?:\.\d+)?\s*(?:kg|g))", re.IGNORECASE)


def extract_weight(text: str):
    """문자열에서 첫 번째 중량 추출 → (중량문자열, 나머지)"""
    m = WEIGHT_PATTERN.search(text)
    if m:
        w = m.group(1).replace(" ", "")
        rest = (text[: m.start()] + text[m.end() :]).strip().strip("/").strip()
        return w, rest
    return "", text


def split_item(raw_name: str):
    """
    품목명_옵션 분리 후 (품목명, 중량, 옵션) 반환
    특수 품목은 별도 처리
    반환값이 리스트인 경우 → 여러 행으로 분리 필요
    """
    # [커피 페스타 1+1] 6월 King콩 ... _옵션/추출원두명(250g)
    # → 행1: King콩 품목  (품목명, 250g, 갈지않음)
    # → 행2: "[커피 페스타 1+1] " + 추출원두명  (250g, 갈지않음)
    if "[커피 페스타 1+1]" in raw_name and "King콩" in raw_name:
        if "_" in raw_name:
            item_part, opt_part = raw_name.split("_", 1)
        else:
            item_part, opt_part = raw_name, ""
        item_name = item_part.strip()
        opt_part = re.sub(r"\(\d*g?\)", "", opt_part).strip()
        if "/" in opt_part:
            before_slash, extra_bean = opt_part.split("/", 1)
            extra_bean = extra_bean.strip()
        else:
            before_slash = opt_part
            extra_bean = ""
        opt1 = before_slash.strip().strip("/").strip()
        row1 = (item_name, "250g", opt1)
        if extra_bean:
            row2 = ("[커피 페스타 1+1] " + extra_bean, "250g", "갈지않음")
            return [row1, row2]
        return row1

    # [커피 페스타 1+1] 액상커피+파우더스틱 ... _옵션1 / 옵션2
    # → 행1: 품목명 그대로, C열 = "/" 앞 옵션
    # → 행2: 품목명 동일, C열 = "/" 뒷 옵션 (수량 동일)
    if "[커피 페스타 1+1]" in raw_name and "액상커피" in raw_name:
        item_name = "[커피 페스타 1+1] 액상커피+파우더스틱"
        opt_part = raw_name.split("_", 1)[1].strip() if "_" in raw_name else ""
        opt_part = re.sub(r"\(\d+개입\)", "", opt_part).strip()
        # 괄호 밖의 첫 번째 "/" 기준으로만 분리
        depth = 0
        slash_idx = -1
        for i, ch in enumerate(opt_part):
            if ch == "(": depth += 1
            elif ch == ")": depth -= 1
            elif ch == "/" and depth == 0:
                slash_idx = i
                break
        if slash_idx != -1:
            opt1 = opt_part[:slash_idx].strip()
            opt2 = opt_part[slash_idx + 1:].strip()
        else:
            opt1, opt2 = opt_part, ""
        row1 = (item_name, "", opt1)
        if opt2:
            row2 = (item_name, "", opt2)
            return [row1, row2]
        return row1

    
    # [첫 구매 찬스] 어센틱 에스프레소 블렌드 250g
    if "[첫 구매 찬스]" in raw_name:
        name = raw_name.replace("[첫 구매 찬스] ", "").replace("250g", "").strip()
        return name, "250g", ""

    # 무료원두 쿠폰 250g
    if "무료원두 쿠폰" in raw_name:
        return "무료원두 쿠폰 250g", "250g", "증정 원두"

    # 이 달의 킹콩 500g  → 처리는 kingkong 단계에서
    if "이 달의 킹콩" in raw_name or "이달의 킹콩" in raw_name:
        return raw_name, "500g", "플러스쿠폰"

    # 이 달의 드립백 (언더스코어 없음)
    if "이 달의 드립백" in raw_name or "이달의 드립백" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip()
            weight, option = extract_weight(option)
            return item_name, weight, option
        return raw_name, "", ""

    # 원두&커피 스쿱 세트 — 갈지않음/(커피명)(250g) 구조
    if "원두&커피 스쿱 세트" in raw_name or "원두 & 커피 스쿱 세트" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip()
            option = option.replace("갈지않음/", "").replace("(250g)", "").strip()
            return item_name, "250g", option
        return raw_name, "250g", ""

    # 일반: 언더스코어로 분리
    if "_" in raw_name:
        parts = raw_name.split("_", 1)
        item_name = parts[0].strip()
        option = parts[1].strip()
        weight, option = extract_weight(option)
        return item_name, weight, option

    # 언더스코어 없음 — 품목명 자체에서 중량 추출 시도
    weight, rest = extract_weight(raw_name)
    if weight:
        return rest if rest else raw_name, weight, ""
    return raw_name, "", ""


# ──────────────────────────────────────────────
# 4. 특수 품목: 킹콩 이름 치환
# ──────────────────────────────────────────────
def resolve_kingkong_name(df: pd.DataFrame) -> pd.DataFrame:
    """'이 달의 킹콩' 행의 품목명을 실제 King콩 품목명으로 교체"""
    king_rows = df[df["품목명"].str.contains(r"[Kk][Ii][Nn][Gg]콩", na=False)]

    mask = df["품목명"].str.contains("이 달의 킹콩|이달의 킹콩", na=False)
    if not mask.any():
        return df

    if not king_rows.empty:
        # 같은 데이터에 King콩 품목이 있으면 그 이름 사용
        king_name = king_rows.iloc[0]["품목명"]
    else:
        # King콩 품목이 없으면 월을 오늘 날짜 기준으로 자동 생성
        from datetime import datetime
        month = datetime.today().month
        king_name = f"[{month}월 KING콩]"

    df.loc[mask, "품목명"] = king_name
    df.loc[mask, "중량"] = "500g"
    df.loc[mask, "옵션"] = "플러스쿠폰"
    return df


# ──────────────────────────────────────────────
# 5. King콩 옵션 정리 (바리스타·농부·농장주 제거 후 합산)
# ──────────────────────────────────────────────
def clean_kingkong_options(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["품목명"].str.contains(r"[Kk][Ii][Nn][Gg]콩", na=False)
    # //와 / 모두 제거 (슬래시 개수 무관)
    for keyword in ["테라로사 바리스타", "에티오피아 농부", "멕시코 농장주"]:
        df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.replace(
            r"\s*/{1,2}\s*" + keyword, "", regex=True
        ).str.strip()
    # 옵션 끝에 남은 슬래시 정리
    df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.strip("/").str.strip()
    return df


# ──────────────────────────────────────────────
# 6. [감사의 달] 옵션 합산 (앞 4자 기준)
# ──────────────────────────────────────────────
def merge_gratitude_month(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["품목명"].str.contains("감사의 달", na=False)
    if not mask.any():
        return df

    gdf = df[mask].copy()
    others = df[~mask].copy()

    merged2_rows = []

    # 옵션이 있는 것과 없는 것 분리
    gdf_with_opt = gdf[gdf["옵션"].str.strip() != ""].copy()
    gdf_no_opt   = gdf[gdf["옵션"].str.strip() == ""].copy()

    # 옵션 있는 것: 앞 5자 기준 합산
    if not gdf_with_opt.empty:
        gdf_with_opt["_key"] = gdf_with_opt["옵션"].str[:5]
        for key, g in gdf_with_opt.groupby("_key", sort=False):
            row = {
                "품목명": "[감사의 달] 2026 선물대전",
                "중량": g.iloc[0]["중량"],
                "옵션": min(g["옵션"].values, key=len),
                "수량": g["수량"].sum(),
            }
            merged2_rows.append(row)

    # 옵션 없는 것: 품목명 그대로 유지, 합산만
    if not gdf_no_opt.empty:
        for name, g in gdf_no_opt.groupby("품목명", sort=False):
            row = {
                "품목명": name,
                "중량": g.iloc[0]["중량"],
                "옵션": "",
                "수량": g["수량"].sum(),
            }
            merged2_rows.append(row)

    merged2 = pd.DataFrame(merged2_rows)
    return pd.concat([others, merged2], ignore_index=True)


# ──────────────────────────────────────────────
# 7. 분류 및 정렬
# ──────────────────────────────────────────────
def classify(row) -> str:
    name = row["품목명"]
    weight = str(row["중량"])
    if "드립백" in name:
        return "드립백"
    if "원두&커피 스쿱 세트" in name or "원두 & 커피 스쿱 세트" in name:
        return "스쿱세트"
    if "세트" in name:
        return "세트"
    if re.search(r"\d+\s*(?:g|kg)", weight, re.IGNORECASE):
        return "원두"
    return "기타"


GROUP_ORDER = {"세트": 0, "기타": 1, "드립백": 2, "스쿱세트": 3, "원두": 4}


def weight_to_gram(w: str) -> float:
    """정렬용 중량→g 변환"""
    w = str(w).strip()
    m = re.match(r"([\d.]+)\s*(kg|g)", w, re.IGNORECASE)
    if not m:
        return 0
    val = float(m.group(1))
    return val * 1000 if m.group(2).lower() == "kg" else val


def aggregate_and_sort(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_group"] = df.apply(classify, axis=1)

    # 그룹별 합산 키 설정
    def agg_key(row):
        g = row["_group"]
        if g in ("세트", "드립백", "스쿱세트"):
            return (row["품목명"], row["옵션"])
        return (row["품목명"], row["중량"], row["옵션"])

    df["_key"] = df.apply(agg_key, axis=1)
    df = df.groupby("_key", sort=False).agg(
        품목명=("품목명", "first"),
        중량=("중량", "first"),
        옵션=("옵션", "first"),
        수량=("수량", "sum"),
        _group=("_group", "first"),
    ).reset_index(drop=True)

    df["_g_order"] = df["_group"].map(GROUP_ORDER)
    df["_w_gram"] = df["중량"].apply(weight_to_gram)

    df.sort_values(
        ["_g_order", "품목명", "_w_gram", "옵션"],
        ascending=[True, True, False, True],
        inplace=True,
    )
    df.reset_index(drop=True, inplace=True)
    return df


# ──────────────────────────────────────────────
# 8. 상품코드 매칭
# ──────────────────────────────────────────────
def match_code(row, code_df: pd.DataFrame) -> str:
    name = str(row["품목명"]).strip()
    weight = str(row["중량"]).strip()
    option = str(row["옵션"]).strip()

    c_code = code_df.columns[0]  # A열 = 상품코드
    c_name = code_df.columns[1]  # B열 = 상품명
    c_opt  = code_df.columns[2]  # C열 = 옵션

    def eq(col, val):
        return code_df[col].str.strip() == val.strip()

    # 1. King콩 (대소문자 무관, 옵션 무관 → 500g으로 매칭)
    if re.search(r"[Kk][Ii][Nn][Gg]콩", name):
        res = code_df[eq(c_name, name) & (code_df[c_opt].str.strip() == "500g")]
        if not res.empty:
            return str(res.iloc[0][c_code])
        # 품목명만으로 fallback
        res = code_df[eq(c_name, name)]
        if not res.empty:
            return str(res.iloc[0][c_code])

    # 2. S.O.S (품목명 + 중량)
    if "S.O.S" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty:
            return str(res.iloc[0][c_code])

    # 3. 원두&커피 스쿱 세트 (옵션에 (250g) 붙여서 매칭)
    if "스쿱 세트" in name or "스쿱세트" in name:
        opt_with_weight = option + "(250g)"
        res = code_df[eq(c_opt, opt_with_weight)]
        if not res.empty:
            return str(res.iloc[0][c_code])

    # 4. TO-GO 세트 (색상 제거 후 매칭)
    if "TO-GO" in name or "to-go" in name.lower():
        opt_no_color = re.sub(r"/블랙|/투명|/화이트|/레드", "", option).strip()
        res = code_df[eq(c_name, name) & eq(c_opt, opt_no_color)]
        if not res.empty:
            return str(res.iloc[0][c_code])

    # 5. 원두 (품목명 + 중량으로 매칭)
    if weight:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty:
            return str(res.iloc[0][c_code])

    # 6. 기본 매칭 (품목명 + 옵션)
    res = code_df[eq(c_name, name) & eq(c_opt, option)]
    if not res.empty:
        return str(res.iloc[0][c_code])

    # 7. 옵션 순서 불일치 (+기준 정렬)
    sorted_opt = "+".join(sorted(option.split("+")))
    name_rows = code_df[code_df[c_name].str.strip() == name]
    if not name_rows.empty:
        match = name_rows[name_rows[c_opt].apply(
            lambda x: "+".join(sorted(str(x).split("+"))) == sorted_opt
        )]
        if not match.empty:
            return str(match.iloc[0][c_code])

    # 8. 품목명만 fallback (옵션 없는 행)
    res = code_df[(code_df[c_name].str.strip() == name) & (code_df[c_opt].str.strip() == "")]
    if not res.empty:
        return str(res.iloc[0][c_code])

    # 9. 옥스포드 피규어 (C열 옵션으로 매칭)
    if "옥스포드" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, option)]
        if not res.empty:
            return str(res.iloc[0][c_code])

    return ""


# ──────────────────────────────────────────────
# 9. 시트3: 바리스타·농부·농장주
# ──────────────────────────────────────────────
def build_sheet3(raw_df: pd.DataFrame) -> pd.DataFrame:
    targets = {
        "테라로사 바리스타": "/ 테라로사 바리스타",
        "에티오피아 농부": "/ 에티오피아 농부",
        "멕시코 농장주": "/ 멕시코 농장주",
    }
    rows = []
    for label, keyword in targets.items():
        mask = raw_df["품목명_원본"].str.contains(keyword, na=False)
        qty = raw_df.loc[mask, "수량"].sum()
        rows.append({
            "품목명": "옥스포드 피규어",
            "빈칸": "",
            "이름": label,
            "수량": qty if qty > 0 else "-",
        })
    return pd.DataFrame(rows)


# ──────────────────────────────────────────────
# 10. 시트2: 원두 중량 합산
# ──────────────────────────────────────────────
def build_sheet2(main_df: pd.DataFrame) -> pd.DataFrame:
    bean_mask = main_df["_group"] == "원두"
    scoop_mask = (main_df["_group"] == "스쿱세트")

    rows = {}

    for _, r in main_df[bean_mask].iterrows():
        name = r["품목명"]
        g = weight_to_gram(r["중량"]) * r["수량"]
        rows[name] = rows.get(name, 0) + g

    for _, r in main_df[scoop_mask].iterrows():
        name = r["옵션"]  # 원두명
        g = 250 * r["수량"]
        rows[name] = rows.get(name, 0) + g

    result = []
    for name, total_g in rows.items():
        result.append({"품목명": name, "중량(kg)": round(total_g / 1000, 3)})
    return pd.DataFrame(result)


# ──────────────────────────────────────────────
# 11. 엑셀 서식 적용
# ──────────────────────────────────────────────
def apply_style(ws, df_with_groups: pd.DataFrame):
    """시트1(주문취합)에 서식 적용"""
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    drip_fill = PatternFill("solid", fgColor=COLOR_DRIP)
    scoop_fill = PatternFill("solid", fgColor=COLOR_SCOOP)
    white_fill = PatternFill("solid", fgColor=COLOR_WHITE)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)

    # 헤더
    headers = ["품목명", "중량", "옵션", "수량", "자사몰상품코드"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row_num = 2
    prev_name = None

    for _, r in df_with_groups.iterrows():
        cur_name = r["품목명"]
        group = r.get("_group", "")

        # 품목명 바뀔 때 빈 행
        if prev_name is not None and cur_name != prev_name:
            for col_idx in range(1, 6):
                ws.cell(row=row_num, column=col_idx).border = THIN_BORDER
            row_num += 1

        # 색상 결정
        if group == "드립백":
            fill = drip_fill
        elif group == "스쿱세트":
            fill = scoop_fill
        else:
            fill = white_fill

        values = [cur_name, r["중량"], r["옵션"], r["수량"], r.get("상품코드", "")]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.border = THIN_BORDER

        prev_name = cur_name
        row_num += 1

    # 열 너비
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def write_simple_sheet(ws, df: pd.DataFrame, title_row: list):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=10)

    for col_idx, h in enumerate(title_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for r_idx, row in df.iterrows():
        for col_idx, val in enumerate(row.values, 1):
            cell = ws.cell(row=r_idx + 2, column=col_idx, value=val)
            cell.font = body_font
            cell.border = THIN_BORDER


# ──────────────────────────────────────────────
# 12. 시트1 상단에 시트3 삽입
# ──────────────────────────────────────────────
def insert_sheet3_into_sheet1(wb: Workbook):
    ws1 = wb["주문취합"]
    ws3 = wb["바리스타·농부·농장주"]

    # 시트3의 2행~마지막 데이터 행 수집
    max_row_3 = ws3.max_row
    insert_count = max_row_3 - 1  # 헤더 제외
    if insert_count <= 0:
        return

    # 시트1의 2행에 삽입 공간 확보
    ws1.insert_rows(2, amount=insert_count + 1)  # +1 빈 행

    # 시트3 데이터 복사 (2행부터)
    for src_row_idx in range(2, max_row_3 + 1):
        dest_row_idx = src_row_idx  # 2→2, 3→3 ...
        for col_idx in range(1, ws3.max_column + 1):
            src_cell = ws3.cell(row=src_row_idx, column=col_idx)
            dst_cell = ws1.cell(row=dest_row_idx, column=col_idx)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)

    # 삽입 후 빈 행 (구분)
    blank_row = 2 + insert_count
    for col_idx in range(1, 6):
        ws1.cell(row=blank_row, column=col_idx).border = THIN_BORDER


# ──────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────
def main(order_file: str, code_file: str, output_file: str = None):
    # 1. 데이터 로드
    raw_df = load_order_data(order_file)
    code_df = load_code_data(code_file)

    # 2. 시트3용 원본 저장 (바리스타·농부·농장주 추출)
    sheet3_df = build_sheet3(raw_df)

    # 3. 텍스트 정리
    raw_df["품목명_정리"] = raw_df["품목명_원본"].apply(clean_item_name)

    # 4. 분리 (품목명, 중량, 옵션) — split_item이 리스트 반환 시 행 확장
    expanded_rows = []
    for _, row in raw_df.iterrows():
        result = split_item(row["품목명_정리"])
        if isinstance(result, list):
            for r in result:
                new_row = row.copy()
                new_row["품목명"] = r[0]
                new_row["중량"] = r[1]
                new_row["옵션"] = r[2]
                expanded_rows.append(new_row)
        else:
            new_row = row.copy()
            new_row["품목명"] = result[0]
            new_row["중량"] = result[1]
            new_row["옵션"] = result[2]
            expanded_rows.append(new_row)
    raw_df = pd.DataFrame(expanded_rows).reset_index(drop=True)

    # 5. S.O.S 중량 보정
    raw_df["중량"] = raw_df.apply(
        lambda r: apply_sos_weight(r["품목명"], r["중량"]), axis=1
    )

    # 6. King콩 이름 치환
    raw_df = resolve_kingkong_name(raw_df)

    # 7. King콩 옵션 정리
    raw_df = clean_kingkong_options(raw_df)

    # 8. [감사의 달] 합산
    raw_df = merge_gratitude_month(raw_df)

    # 9. 분류·합산·정렬
    main_df = aggregate_and_sort(raw_df)

    # 10. 상품코드 매칭
    main_df["상품코드"] = main_df.apply(lambda r: match_code(r, code_df), axis=1)

    # 11. 시트2: 원두 중량 합산
    sheet2_df = build_sheet2(main_df)

    # 12. 엑셀 출력
    wb = Workbook()

    # 시트1: 주문취합
    ws1 = wb.active
    ws1.title = "주문취합"
    apply_style(ws1, main_df)

    # 시트2: 원두 중량 합산
    ws2 = wb.create_sheet("원두 중량 합산")
    write_simple_sheet(ws2, sheet2_df, ["품목명", "중량(kg)"])

    # 시트3: 바리스타·농부·농장주
    ws3 = wb.create_sheet("바리스타·농부·농장주")
    write_simple_sheet(ws3, sheet3_df, ["품목명", "빈칸", "이름", "수량"])

    # 13. 시트3 데이터를 시트1 상단에 삽입
    insert_sheet3_into_sheet1(wb)

    # 14. 저장
    today = datetime.today().strftime("%Y%m%d")
    output_path = output_file if output_file else f"자사몰주문취합_{today}.xlsx"
    wb.save(output_path)
    print(f"저장 완료: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("사용법: python terrarosa_order_processor.py <주문취합.xlsx> <자사몰상품코드.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
