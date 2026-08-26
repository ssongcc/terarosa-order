"""
테라로사 자사몰 주문취합 + 분배모드 Streamlit 앱
실행: streamlit run app.py
"""

import json
import re
import os
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook

# GitHub 영구 저장소
try:
    from github_storage import gh_load, gh_save
    _USE_GITHUB = True
except Exception:
    _USE_GITHUB = False
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 분배모드 로직
from das_distribution import run as das_run

# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────
CONFIG_PATH = Path("set_config.json")
# ── 첫 구매 찬스 상품 매핑 (매달 변경 시 여기만 수정) ──────────────────
FIRST_PURCHASE_ITEM = {
    "name":   "하우스 드립 블렌드",  # A열 품목명 ([첫 구매 찬스] 제거)
    "weight": "250g",               # B열 중량
    "option": "갈지않음",           # C열 옵션
}
# ────────────────────────────────────────────────────────────────────────

# ── 품목명 앞 접두사 제거 목록 (추가/삭제 시 여기만 수정) ─────────────────
# 제거 후 기존 상품과 합산, 원두인 경우 중량 합산 시트에도 포함
PREFIX_REMOVE = [
    "[테라로사 BEST 8] ",   # 예: "[테라로사 BEST 8] 강릉 블렌드" → "강릉 블렌드"
    # 추가 예시: "[기획전] ",
]
# ────────────────────────────────────────────────────────────────────────

# ── 세트이지만 원두로 분류할 품목 키워드 (추가/삭제 시 여기만 수정) ────────
# 품목명에 "세트"가 있어도 원두로 분류하여 중량 합산에 포함
BEAN_SET_KEYWORDS = [
    "보름달 블렌드 & 드리퍼 세트",   # 26년 추석
    # "추가 예시 세트명",
]
# ────────────────────────────────────────────────────────────────────────

# ── 세트 구성품 자동 추가 규칙 (추가/삭제 시 여기만 수정) ────────────────
# 품목명 키워드 매칭 시 원본 행 유지 + 추가 행(add_rows) 삽입
# add_rows: [{"name": A열, "weight": B열, "option": C열, "qty_mult": 수량배수}, ...]
BONUS_ITEM_RULES = [
    {
        "keyword": "보름달 블렌드 & 드리퍼 세트",
        "add_rows": [
            {"name": "보름달 블렌드 & 드리퍼 세트", "weight": "", "option": "테라로사 드리퍼", "qty_mult": 1},
        ],
    },
    {
        "keyword": "엑스트라 버진 올리브 오일 세트",
        "add_rows": [
            {"name": "[26년 추석] 엑스트라 버진 올리브 오일 세트", "weight": "", "option": "미틸로 오가닉 올리브오일 250ml", "qty_mult": 2},
        ],
    },
    {
        "keyword": "홈브루잉 스타터팩",
        "replace_original": True,   # 원본 행 대신 아래 행들로 완전 교체
        "add_rows": [
            {"name": "[26년 추석] 홈브루잉 스타터팩",         "weight": "",     "option": "드리퍼",   "qty_mult": 1},
            {"name": "어센틱 에스프레소 블렌드",               "weight": "250g", "option": "갈지않음", "qty_mult": 1},
            {"name": "에티오피아 예가체페 아리차 토착종 워시드", "weight": "250g", "option": "갈지않음", "qty_mult": 1},
        ],
    },
    # 추가 예시:
    # {"keyword": "새 세트명", "add_rows": [{"name": "구성품A", "weight": "250g", "option": "갈지않음", "qty_mult": 1}]},
]
# ────────────────────────────────────────────────────────────────────────

# ── 추석 세트 상품 목록 (시즌 종료 시 빈 리스트로 변경) ───────────────────
CHUSEOK_SETS = [
    "[26년 추석] 간편커피 & 머그 세트",
    "[26년 추석] 드립백 & 머그 세트",
    "[26년 추석] 싱글오리진 파우더 커피 세트 (30개입)",
    "[26년 추석] 엑스트라 버진 올리브 오일 세트",
    "[26년 추석] 보름달 블렌드 & 드리퍼 세트",
    "[26년 추석] 홈브루잉 스타터팩",
]
# ────────────────────────────────────────────────────────────────────────

# ── 드립백 증정 규칙 (변경 시 여기만 수정) ──────────────────────────────
DRIP_GIFT_RULES = [
    # (품목명 포함 키워드, 옵션 포함 키워드(없으면 None), 증정 품목명, 수량 배수)
    ("드립백 (100개)",  None,      "[증정] 에티오피아 드립백 10개", 1),
    ("드립백 (200개)",  None,      "[증정] 에티오피아 드립백 10개", 2),
    ("드립백 정기 배송", "100개입", "[증정] 에티오피아 드립백 10개", 1),
]
# ────────────────────────────────────────────────────────────────────────

REMOVE_STRINGS = [
    "/구매 안함", "/플러스", "[플러스] ", "/불필요", "/필요",
    "불필요", "필요", "/상자 없음", "/포장 없음",
    "테라로사 시그니처 ", "[Online Exclusive] ", "[Online Exclusive/플러스] ",
    "[C.O.E/플러스] ",
]
TEXT_REPLACE = {
    "중간 분쇄(드립용)": "드립용",
    "가는 분쇄(에스프레소용)": "에스프레소용",
}
COLOR_DRIP   = "E2EFDA"
COLOR_SCOOP  = "FFF2CC"
COLOR_HEADER = "D9D9D9"
COLOR_WHITE  = "FFFFFF"
COLOR_SET    = "DDEEFF"
COL_WIDTHS   = {"A": 42, "B": 10, "C": 35, "D": 8, "E": 18}
THIN_BORDER  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WEIGHT_PATTERN = re.compile(r"(\d+(?:\.\d+)?\s*(?:kg|g))", re.IGNORECASE)

# ──────────────────────────────────────────────
# 설정 로드/저장
# ──────────────────────────────────────────────
def load_set_config() -> dict:
    # 로컬 파일 우선 (Cloud/로컬 공통)
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    if _USE_GITHUB:
        try:
            return gh_load("set_config.json", {})
        except Exception:
            pass
    return {}

def save_set_config(cfg: dict):
    # 로컬 파일에 항상 저장
    try:
        CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    # GitHub에도 저장 (Cloud 환경)
    if _USE_GITHUB:
        try:
            gh_save("set_config.json", cfg)
        except Exception:
            pass

# ──────────────────────────────────────────────
# 세트/쿠폰 로직 (기존 그대로)
# ──────────────────────────────────────────────
def expand_drip_gift(df):
    """드립백 100/200개, 정기 배송 100개입 → 에티오피아 드립백 증정 행 추가"""
    expanded = []
    for _, row in df.iterrows():
        expanded.append(row)
        name = str(row.get("품목명", "") or "").strip()
        opt  = str(row.get("옵션", "") or "").strip()
        qty  = int(row.get("수량", 1) or 1)
        for kw_name, kw_opt, gift_name, mult in DRIP_GIFT_RULES:
            if kw_name in name:
                if kw_opt is None or kw_opt in opt:
                    gift_row = row.copy()
                    gift_row["품목명"] = gift_name
                    gift_row["중량"]   = ""
                    gift_row["옵션"]   = ""
                    gift_row["수량"]   = qty * mult
                    expanded.append(gift_row)
                    break
    return pd.DataFrame(expanded).reset_index(drop=True)

def expand_bonus_items(df):
    """세트 구성품 자동 추가 (BONUS_ITEM_RULES 기준) — 다중 행 추가 지원
    replace_original=True인 규칙은 원본 행을 add_rows로 완전 대체
    replace_original=False(기본)는 원본 행 유지 + add_rows 추가
    """
    expanded = []
    for _, row in df.iterrows():
        name = str(row.get("품목명", "") or "").strip()
        qty  = int(row.get("수량", 1) or 1)
        matched = False
        for rule in BONUS_ITEM_RULES:
            if rule["keyword"] in name:
                matched = True
                if not rule.get("replace_original", False):
                    expanded.append(row)  # 원본 유지
                for add in rule["add_rows"]:
                    bonus = row.copy()
                    bonus["품목명"] = add["name"]
                    bonus["중량"]   = add["weight"]
                    bonus["옵션"]   = add["option"]
                    bonus["수량"]   = qty * add["qty_mult"]
                    expanded.append(bonus)
                break
        if not matched:
            expanded.append(row)
    return pd.DataFrame(expanded).reset_index(drop=True)

def expand_set_items(df, set_config):
    if not set_config:
        return df
    expanded = []
    for _, row in df.iterrows():
        name = str(row.get("품목명", "")).strip()
        matched = None
        for set_name, components in set_config.items():
            if set_name in name:
                matched = (set_name, components)
                break
        if matched and matched[1]:
            set_name, components = matched
            qty = int(row.get("수량", 1))
            for comp in components:
                new_row = row.copy()
                new_row["품목명_원본"] = f"[세트분리] {set_name} → {comp['name']}"
                new_row["품목명"]  = set_name
                new_row["중량"]    = comp.get("weight", "")
                new_row["옵션"]    = comp["name"]
                new_row["수량"]    = qty * comp.get("qty", 1)
                new_row["_is_set_expanded"] = True
                expanded.append(new_row)
        else:
            row = row.copy()
            if "_is_set_expanded" not in row.index:
                row["_is_set_expanded"] = False
            expanded.append(row)
    return pd.DataFrame(expanded).reset_index(drop=True)

def load_coupon_config():
    if _USE_GITHUB:
        return gh_load("coupon_config.json", [])
    coupon_path = Path("coupon_config.json")
    if coupon_path.exists():
        try:
            return json.loads(coupon_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []

def expand_coupon_items(df, coupon_config):
    if not coupon_config:
        return df
    expanded = []
    for _, row in df.iterrows():
        name = str(row.get("품목명", "")).strip()
        if "무료원두 쿠폰" in name:
            for comp in coupon_config:
                if not comp.get("name"):
                    continue
                new_row = row.copy()
                new_row["품목명"]  = comp["name"]
                new_row["중량"]    = comp.get("weight", "250g")
                new_row["옵션"]    = comp.get("option", "증정 원두")
                new_row["수량"]    = int(comp.get("qty", 1))
                expanded.append(new_row)
        else:
            expanded.append(row)
    return pd.DataFrame(expanded).reset_index(drop=True)

# ──────────────────────────────────────────────
# 주문취합 처리 로직 (기존 그대로)
# ──────────────────────────────────────────────
def load_order_data(file):
    df = pd.read_excel(file, sheet_name="취합용", header=0, dtype=str)
    df = df.iloc[:, :2].copy()
    df.columns = ["품목명_원본", "수량"]
    df.dropna(subset=["품목명_원본"], inplace=True)
    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    return df.reset_index(drop=True)

def load_code_data(file):
    df = pd.read_excel(file, header=0, dtype=str)
    df.fillna("", inplace=True)
    return df

def clean_item_name(name):
    for s in REMOVE_STRINGS:
        name = name.replace(s, "")
    for prefix in PREFIX_REMOVE:
        if name.startswith(prefix):
            name = name[len(prefix):]
    for old, new in TEXT_REPLACE.items():
        name = name.replace(old, new)
    if "어센틱" in name and "정기 배송" in name:
        if "_" in name:
            option = name.split("_", 1)[1]
            name = "어센틱 에스프레소 블렌드_" + option
        else:
            name = "어센틱 에스프레소 블렌드"
    return name.strip()

def apply_sos_weight(name, weight):
    if "S.O.S" in name and weight == "300g":
        return "150g"
    return weight

def extract_weight(text):
    m = WEIGHT_PATTERN.search(text)
    if m:
        w = m.group(1).replace(" ", "")
        rest = (text[:m.start()] + text[m.end():]).strip().strip("/").strip()
        return w, rest
    return "", text

def split_item(raw_name):
    if "[커피 페스타 1+1]" in raw_name and ("KING콩" in raw_name or "King콩" in raw_name):
        item_part = raw_name.split("_", 1)[0].strip() if "_" in raw_name else raw_name
        item_part = re.sub(r"\+(블렌드|싱글오리진)$", "", item_part).strip()
        opt_raw = raw_name.split("_", 1)[1] if "_" in raw_name else ""
        for old, new in TEXT_REPLACE.items():
            opt_raw = opt_raw.replace(old, new)
        if "/" in opt_raw:
            grind_opt, gift_bean = opt_raw.split("/", 1)
            grind_opt = grind_opt.strip()
            gift_bean = gift_bean.strip()
        else:
            grind_opt = opt_raw.strip()
            gift_bean = ""
        row1 = (item_part, "250g", grind_opt)
        if gift_bean:
            return [row1, ("[커피 페스타 증정] " + gift_bean, "250g", "갈지않음")]
        return row1

    if "[커피 페스타 1+1]" in raw_name and "액상커피" in raw_name:
        item_name_orig = "[커피 페스타 1+1] 액상커피+파우더스틱"
        item_name_gift = "[커피 페스타 증정] 액상커피+파우더스틱"
        parts = raw_name.split("_")
        opt_part = parts[1].strip() if len(parts) >= 2 else ""
        opt_part = re.sub(r"\(\d+개입\)", "", opt_part).strip()
        depth, plus_idx = 0, -1
        for i, ch in enumerate(opt_part):
            if ch == "(": depth += 1
            elif ch == ")": depth -= 1
            elif ch == "+" and depth == 0:
                plus_idx = i; break
        if plus_idx != -1:
            opt1 = opt_part[:plus_idx].strip()
            opt2 = opt_part[plus_idx+1:].strip()
        else:
            opt1, opt2 = opt_part, ""
        row1 = (item_name_orig, "", opt1)
        if opt2:
            return [row1, (item_name_gift, "", opt2)]
        return row1

    if "[첫 구매 찬스]" in raw_name:
        return FIRST_PURCHASE_ITEM["name"], FIRST_PURCHASE_ITEM["weight"], FIRST_PURCHASE_ITEM["option"]

    if "무료원두 쿠폰" in raw_name:
        return "무료원두 쿠폰 250g", "250g", "증정 원두"

    if "이 달의 킹콩 500g" in raw_name or "이달의 킹콩 500g" in raw_name:
        return "[8월 KING콩] 인도네시아 아체 리방가요 P88 워시드", "500g", "플러스쿠폰"

    if "이 달의 드립백" in raw_name or "이달의 드립백" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip()
            weight, option = extract_weight(option)
            return item_name, weight, option
        return raw_name, "", ""

    if "원두&커피 스쿱 세트" in raw_name or "원두 & 커피 스쿱 세트" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip().replace("갈지않음/", "").replace("(250g)", "").strip()
            return item_name, "250g", option
        return raw_name, "250g", ""

    if "_" in raw_name:
        parts = raw_name.split("_", 1)
        item_name = parts[0].strip()
        option = parts[1].strip()
        weight, option = extract_weight(option)
        return item_name, weight, option

    weight, rest = extract_weight(raw_name)
    if weight:
        return rest if rest else raw_name, weight, ""
    return raw_name, "", ""

def resolve_kingkong_name(df):
    return df

def clean_kingkong_options(df):
    mask = df["품목명"].str.contains(r"[Kk][Ii][Nn][Gg]콩", na=False)
    for keyword in ["테라로사 바리스타", "에티오피아 농부", "멕시코 농장주"]:
        df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.replace(
            r"\s*/{1,2}\s*" + keyword, "", regex=True
        ).str.strip()
    df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.strip("/").str.strip()
    return df

def merge_gratitude_month(df):
    mask = df["품목명"].str.contains("감사의 달", na=False)
    if not mask.any():
        return df
    gdf, others = df[mask].copy(), df[~mask].copy()
    merged_rows = []
    gdf_with = gdf[gdf["옵션"].str.strip() != ""].copy()
    gdf_no   = gdf[gdf["옵션"].str.strip() == ""].copy()
    if not gdf_with.empty:
        gdf_with["_key"] = gdf_with["옵션"].str[:5]
        for key, g in gdf_with.groupby("_key", sort=False):
            merged_rows.append({"품목명": "[감사의 달] 2026 선물대전", "중량": g.iloc[0]["중량"],
                                 "옵션": min(g["옵션"].values, key=len), "수량": g["수량"].sum()})
    if not gdf_no.empty:
        for name, g in gdf_no.groupby("품목명", sort=False):
            merged_rows.append({"품목명": name, "중량": g.iloc[0]["중량"], "옵션": "", "수량": g["수량"].sum()})
    return pd.concat([others, pd.DataFrame(merged_rows)], ignore_index=True)

def classify(row):
    name, weight = row["품목명"], str(row["중량"])
    if "드립백" in name: return "드립백"
    if "원두&커피 스쿱 세트" in name or "원두 & 커피 스쿱 세트" in name: return "스쿱세트"
    # BEAN_SET_KEYWORDS에 있으면 세트라도 원두로 분류
    if any(kw in name for kw in BEAN_SET_KEYWORDS):
        if re.search(r"\d+\s*(?:g|kg)", weight, re.IGNORECASE): return "원두"
    if "세트" in name: return "세트"
    if re.search(r"\d+\s*(?:g|kg)", weight, re.IGNORECASE): return "원두"
    return "기타"

GROUP_ORDER = {"세트": 0, "기타": 1, "드립백": 2, "스쿱세트": 3, "원두": 4}

def weight_to_gram(w):
    w = str(w).strip()
    m = re.match(r"([\d.]+)\s*(kg|g)", w, re.IGNORECASE)
    if not m: return 0
    val = float(m.group(1))
    return val * 1000 if m.group(2).lower() == "kg" else val

def aggregate_and_sort(df):
    df = df.copy()
    df["_group"] = df.apply(classify, axis=1)
    def agg_key(row):
        g = row["_group"]
        if g in ("세트", "드립백", "스쿱세트"):
            return (row["품목명"], row["옵션"])
        return (row["품목명"], row["중량"], row["옵션"])
    df["_key"] = df.apply(agg_key, axis=1)
    df = df.groupby("_key", sort=False).agg(
        품목명=("품목명", "first"), 중량=("중량", "first"), 옵션=("옵션", "first"),
        수량=("수량", "sum"), _group=("_group", "first"),
    ).reset_index(drop=True)
    df["_g_order"] = df["_group"].map(GROUP_ORDER)
    df["_w_gram"]  = df["중량"].apply(weight_to_gram)
    df.sort_values(["_g_order", "품목명", "_w_gram", "옵션"],
                   ascending=[True, True, False, True], inplace=True)
    return df.reset_index(drop=True)

def match_code(row, code_df):
    name   = str(row["품목명"]).strip()
    weight = str(row["중량"]).strip()
    option = str(row["옵션"]).strip()
    c_code, c_name, c_opt = code_df.columns[0], code_df.columns[1], code_df.columns[2]
    def eq(col, val): return code_df[col].str.strip() == val.strip()

    if re.search(r"[Kk][Ii][Nn][Gg]콩", name):
        res = code_df[eq(c_name, name) & (code_df[c_opt].str.strip() == "500g")]
        if not res.empty: return str(res.iloc[0][c_code])
        res = code_df[eq(c_name, name)]
        if not res.empty: return str(res.iloc[0][c_code])
    if "S.O.S" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty: return str(res.iloc[0][c_code])
    if "스쿱 세트" in name or "스쿱세트" in name:
        res = code_df[eq(c_opt, option + "(250g)")]
        if not res.empty: return str(res.iloc[0][c_code])
    if "TO-GO" in name or "to-go" in name.lower():
        opt_no_color = re.sub(r"/블랙|/투명|/화이트|/레드", "", option).strip()
        res = code_df[eq(c_name, name) & eq(c_opt, opt_no_color)]
        if not res.empty: return str(res.iloc[0][c_code])
    if weight:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty: return str(res.iloc[0][c_code])
    res = code_df[eq(c_name, name) & eq(c_opt, option)]
    if not res.empty: return str(res.iloc[0][c_code])
    sorted_opt = "+".join(sorted(option.split("+")))
    name_rows  = code_df[code_df[c_name].str.strip() == name]
    if not name_rows.empty:
        match = name_rows[name_rows[c_opt].apply(
            lambda x: "+".join(sorted(str(x).split("+"))) == sorted_opt)]
        if not match.empty: return str(match.iloc[0][c_code])
    res = code_df[(code_df[c_name].str.strip() == name) & (code_df[c_opt].str.strip() == "")]
    if not res.empty: return str(res.iloc[0][c_code])
    if "옥스포드" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, option)]
        if not res.empty: return str(res.iloc[0][c_code])
    return ""

def build_sheet3(raw_df):
    targets = {"테라로사 바리스타": "/ 테라로사 바리스타",
               "에티오피아 농부": "/ 에티오피아 농부",
               "멕시코 농장주": "/ 멕시코 농장주"}
    rows = []
    for label, keyword in targets.items():
        mask = raw_df["품목명_원본"].str.contains(keyword, na=False)
        qty  = raw_df.loc[mask, "수량"].sum()
        if qty > 0:
            rows.append({"품목명": "옥스포드 피규어", "빈칸": "", "이름": label, "수량": qty})
    return pd.DataFrame(rows)

def build_sheet_chuseok(raw_df_original):
    """추석 세트 상품 합산 시트 생성"""
    if not CHUSEOK_SETS:
        return pd.DataFrame()
    rows = []
    for set_name in CHUSEOK_SETS:
        # 품목명_원본에서 해당 세트명 포함 행 찾아 수량 합산
        mask = raw_df_original["품목명_원본"].astype(str).str.contains(
            re.escape(set_name.split("]")[1].strip()) if "]" in set_name else re.escape(set_name),
            regex=True, na=False
        )
        qty = int(raw_df_original.loc[mask, "수량"].sum()) if mask.any() else 0
        rows.append({"세트명": set_name, "수량": qty})
    return pd.DataFrame(rows)

def build_sheet_best8(raw_df_before_prefix):
    """[테라로사 BEST 8] 상품만 추출 - 접두사 제거 전 원본 기준"""
    mask = raw_df_before_prefix["품목명_원본"].astype(str).str.contains(
        r"\[테라로사 BEST 8\]", regex=True, na=False
    )
    df = raw_df_before_prefix[mask].copy()
    if df.empty:
        return pd.DataFrame()
    # 접두사 제거 후 기존 split_item 로직으로 분해
    rows = []
    for _, row in df.iterrows():
        cleaned = str(row["품목명_원본"]).strip()
        for prefix in PREFIX_REMOVE:
            if cleaned.startswith(prefix):
                cleaned = cleaned[len(prefix):]
        result = split_item(cleaned)
        if isinstance(result, list):
            for r in result:
                rows.append({"품목명": r[0], "중량": r[1], "옵션": r[2], "수량": row["수량"]})
        else:
            rows.append({"품목명": result[0], "중량": result[1], "옵션": result[2], "수량": row["수량"]})
    df_out = pd.DataFrame(rows)
    # 동일 품목 합산 후 주문취합 동일 정렬
    df_out["수량"] = pd.to_numeric(df_out["수량"], errors="coerce").fillna(0).astype(int)
    df_out = df_out.groupby(["품목명","중량","옵션"], sort=False, as_index=False)["수량"].sum()
    df_out["_group"]   = df_out.apply(classify, axis=1)
    df_out["_g_order"] = df_out["_group"].map(GROUP_ORDER)
    df_out["_w_gram"]  = df_out["중량"].apply(weight_to_gram)
    df_out = df_out.sort_values(["_g_order","품목명","_w_gram","옵션"],
                                ascending=[True,True,False,True])
    return df_out.drop(columns=["_group","_g_order","_w_gram"]).reset_index(drop=True)

def build_sheet2(main_df):
    rows = {}
    king_mask = (
        main_df["품목명"].str.contains(r"\[커피 페스타 1\+1\]", regex=True, na=False) &
        main_df["품목명"].str.contains("KING콩|King콩", na=False)
    )
    for _, r in main_df[(main_df["_group"] == "원두") & ~king_mask].iterrows():
        name = re.sub(r"^\[커피 페스타 증정\]\s*", "", r["품목명"]).strip()
        rows[name] = rows.get(name, 0) + weight_to_gram(r["중량"]) * r["수량"]
    for _, r in main_df[main_df["_group"] == "스쿱세트"].iterrows():
        name = r["옵션"]
        rows[name] = rows.get(name, 0) + 250 * r["수량"]
    for _, r in main_df[king_mask].iterrows():
        name = re.sub(r"^\[커피 페스타 1\+1\]\s*", "", r["품목명"]).strip()
        rows[name] = rows.get(name, 0) + 250 * r["수량"]
    return pd.DataFrame([{"품목명": n, "중량(kg)": round(g / 1000, 3)} for n, g in rows.items()])

def apply_style(ws, df_with_groups):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    drip_fill   = PatternFill("solid", fgColor=COLOR_DRIP)
    scoop_fill  = PatternFill("solid", fgColor=COLOR_SCOOP)
    white_fill  = PatternFill("solid", fgColor=COLOR_WHITE)
    set_fill    = PatternFill("solid", fgColor=COLOR_SET)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font   = Font(name="Arial", size=10)
    headers = ["품목명", "중량", "옵션", "수량", "자사몰상품코드"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = THIN_BORDER
    row_num, prev_name = 2, None
    for _, r in df_with_groups.iterrows():
        cur_name = r["품목명"]
        group    = r.get("_group", "")
        is_set   = r.get("_is_set_expanded", False)
        if prev_name is not None and cur_name != prev_name:
            for col_idx in range(1, 6):
                ws.cell(row=row_num, column=col_idx).border = THIN_BORDER
            row_num += 1
        if is_set:          fill = set_fill
        elif group == "드립백":  fill = drip_fill
        elif group == "스쿱세트": fill = scoop_fill
        else:               fill = white_fill
        values = [cur_name, r["중량"], r["옵션"], r["수량"], r.get("상품코드", "")]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = body_font; cell.fill = fill; cell.border = THIN_BORDER
        prev_name = cur_name
        row_num += 1
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def write_simple_sheet(ws, df, title_row):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font   = Font(name="Arial", size=10)
    for col_idx, h in enumerate(title_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = THIN_BORDER
    for r_idx, row in df.iterrows():
        for col_idx, val in enumerate(row.values, 1):
            cell = ws.cell(row=r_idx + 2, column=col_idx, value=val)
            cell.font = body_font; cell.border = THIN_BORDER

def insert_sheet3_into_sheet1(wb):
    ws1, ws3 = wb["주문취합"], wb["바리스타·농부·농장주"]
    max_row_3   = ws3.max_row
    insert_count = max_row_3 - 1
    if insert_count <= 0: return
    ws1.insert_rows(2, amount=insert_count + 1)
    for src_row_idx in range(2, max_row_3 + 1):
        dest_row_idx = src_row_idx
        for col_idx in range(1, ws3.max_column + 1):
            src_cell = ws3.cell(row=src_row_idx, column=col_idx)
            dst_cell = ws1.cell(row=dest_row_idx, column=col_idx)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font); dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border); dst_cell.alignment = copy(src_cell.alignment)
    blank_row = 2 + insert_count
    for col_idx in range(1, 6):
        ws1.cell(row=blank_row, column=col_idx).border = THIN_BORDER

def postprocess_festa_rows(ws):
    FILL_FESTA = PatternFill("solid", fgColor="DDEEFF")
    FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
    FILL_BLANK = PatternFill("solid", fgColor="FFFFFF")
    for row in ws.iter_rows():
        a_val = row[0].value
        if a_val and "커피 페스타" in str(a_val):
            for cell in row[:5]:
                cell.fill = FILL_FESTA

    def insert_blank(ws, row_idx):
        ws.insert_rows(row_idx)
        for col in range(1, 6):
            c = ws.cell(row=row_idx, column=col)
            c.value = None; c.fill = FILL_BLANK; c.border = THIN_BORDER

    def is_king_brazil(a):
        return a and "[커피 페스타 1+1]" in str(a) and "KING콩" in str(a) and "브라질" in str(a)
    def is_king_ethiopia(a):
        return a and "[커피 페스타 1+1]" in str(a) and "KING콩" in str(a) and "에티오피아" in str(a)
    def is_gift_bean(a, b):
        return a and "[커피 페스타 증정]" in str(a) and str(b) == "250g"
    def is_festa_1p1_liquid(a):
        return a and "[커피 페스타 1+1]" in str(a) and "액상" in str(a)
    def is_gift_liquid(a, b):
        return a and "[커피 페스타 증정]" in str(a) and str(b) != "250g"
    def is_festa_item(a):
        return a and "[커피 페스타]" in str(a)

    changed = True
    while changed:
        changed = False
        rows_data = list(ws.iter_rows(values_only=True))
        for i in range(1, len(rows_data)):
            a = rows_data[i][0]
            if a is not None: continue
            prev_a = rows_data[i-1][0]
            next_a = rows_data[i+1][0] if i+1 < len(rows_data) else None
            prev_b = rows_data[i-1][1]
            next_b = rows_data[i+1][1] if i+1 < len(rows_data) else None
            if is_gift_bean(prev_a, prev_b) and is_gift_bean(next_a, next_b):
                ws.delete_rows(i + 1); changed = True; break

    rows_data = list(ws.iter_rows(values_only=True))
    to_insert = []
    for i in range(1, len(rows_data)):
        prev_a = rows_data[i-1][0]; curr_a = rows_data[i][0]
        prev_b = rows_data[i-1][1]; curr_b = rows_data[i][1]
        if prev_a is None or curr_a is None: continue
        if is_king_brazil(prev_a) and is_king_ethiopia(curr_a): to_insert.append(i + 1)
        elif is_king_ethiopia(prev_a) and is_gift_bean(curr_a, curr_b): to_insert.append(i + 1)
        elif is_gift_bean(prev_a, prev_b) and is_festa_1p1_liquid(curr_a): to_insert.append(i + 1)
        elif is_festa_1p1_liquid(prev_a) and is_gift_liquid(curr_a, curr_b): to_insert.append(i + 1)
        elif is_gift_liquid(prev_a, prev_b) and not is_gift_liquid(curr_a, curr_b): to_insert.append(i + 1)
        elif is_festa_item(prev_a) and is_festa_item(curr_a) and str(prev_a) != str(curr_a): to_insert.append(i + 1)

    for idx in sorted(set(to_insert), reverse=True):
        insert_blank(ws, idx)

    for row in ws.iter_rows():
        a_val = row[0].value
        if a_val and "커피 페스타" in str(a_val):
            for cell in row[:5]:
                cell.fill = FILL_FESTA
    for row in ws.iter_rows():
        c_val = row[2].value
        if c_val and "플러스쿠폰" in str(c_val):
            for cell in row[:5]:
                cell.fill = FILL_GREEN

def process(order_file, code_file, set_config):
    raw_df  = load_order_data(order_file)
    code_df = load_code_data(code_file)
    sheet3_df = build_sheet3(raw_df)
    # 접두사 제거 전 원본 캡처 (BEST8, 추석 시트용)
    raw_df_for_best8 = raw_df.copy()
    raw_df_original  = raw_df.copy()
    raw_df["품목명_정리"] = raw_df["품목명_원본"].apply(clean_item_name)
    expanded_rows = []
    for _, row in raw_df.iterrows():
        result = split_item(row["품목명_정리"])
        if isinstance(result, list):
            for r in result:
                new_row = row.copy(); new_row["품목명"] = r[0]
                new_row["중량"] = r[1]; new_row["옵션"] = r[2]
                expanded_rows.append(new_row)
        else:
            new_row = row.copy(); new_row["품목명"] = result[0]
            new_row["중량"] = result[1]; new_row["옵션"] = result[2]
            expanded_rows.append(new_row)
    raw_df = pd.DataFrame(expanded_rows).reset_index(drop=True)
    raw_df["중량"] = raw_df.apply(lambda r: apply_sos_weight(r["품목명"], r["중량"]), axis=1)
    raw_df = resolve_kingkong_name(raw_df)
    raw_df = clean_kingkong_options(raw_df)
    raw_df = merge_gratitude_month(raw_df)
    coupon_config = load_coupon_config()
    raw_df = expand_coupon_items(raw_df, coupon_config)
    raw_df = expand_drip_gift(raw_df)
    raw_df = expand_bonus_items(raw_df)
    raw_df = expand_set_items(raw_df, set_config)
    main_df = aggregate_and_sort(raw_df)
    if "_is_set_expanded" in raw_df.columns:
        set_flags = raw_df.groupby(
            raw_df.apply(lambda r: (r["품목명"], r.get("중량",""), r.get("옵션","")), axis=1)
        )["_is_set_expanded"].first()
        def get_flag(r):
            try: return set_flags[(r["품목명"], r["중량"], r["옵션"])]
            except: return False
        main_df["_is_set_expanded"] = main_df.apply(get_flag, axis=1)
    main_df["상품코드"] = main_df.apply(lambda r: match_code(r, code_df), axis=1)
    sheet2_df = build_sheet2(main_df)
    wb  = Workbook()
    ws1 = wb.active; ws1.title = "주문취합"
    apply_style(ws1, main_df)
    ws2 = wb.create_sheet("원두 중량 합산")
    write_simple_sheet(ws2, sheet2_df, ["품목명", "중량(kg)"])
    if not sheet3_df.empty:
        ws3 = wb.create_sheet("바리스타·농부·농장주")
    # 추석 세트 현황 시트
    chuseok_df = build_sheet_chuseok(raw_df_original)
    if not chuseok_df.empty and CHUSEOK_SETS:
        ws_chu = wb.create_sheet("추석 세트 현황")
        ws_chu.append(["세트명", "수량"])
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="8B3A2A")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        body_font   = Font(name="Arial", size=10)
        for c in ws_chu[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center"); c.border = THIN_BORDER
        for _, r in chuseok_df.iterrows():
            ws_chu.append([r["세트명"], r["수량"]])
        for r in range(2, ws_chu.max_row + 1):
            for c in ws_chu[r]:
                c.font = body_font; c.border = THIN_BORDER
        # 합계 행
        last = ws_chu.max_row + 1
        ws_chu.cell(last, 1, "합계").font = Font(name="Arial", size=10, bold=True)
        ws_chu.cell(last, 2, f"=SUM(B2:B{last-1})").font = Font(name="Arial", size=10, bold=True)
        for c in ws_chu[last]:
            c.fill = PatternFill("solid", fgColor=COLOR_HEADER); c.border = THIN_BORDER
        ws_chu.column_dimensions["A"].width = 45
        ws_chu.column_dimensions["B"].width = 8

    best8_df = build_sheet_best8(raw_df_for_best8)
    if not best8_df.empty:
        ws_best8 = wb.create_sheet("BEST8 주문 현황")
        write_simple_sheet(ws_best8, best8_df.rename(columns={
            "품목명":"품목명","중량":"중량","옵션":"옵션","수량":"수량"}),
            ["품목명","중량","옵션","수량"])
        # 주문취합과 동일 스타일 적용
        ws_best8.column_dimensions["A"].width = 42
        ws_best8.column_dimensions["B"].width = 10
        ws_best8.column_dimensions["C"].width = 30
        ws_best8.column_dimensions["D"].width = 8
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        header_font = Font(name="Arial", size=10, bold=True)
        body_font   = Font(name="Arial", size=10)
        white_fill  = PatternFill("solid", fgColor=COLOR_WHITE)
        drip_fill   = PatternFill("solid", fgColor=COLOR_DRIP)
        for c in ws_best8[1]:
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal="center"); c.border = THIN_BORDER
        prev_name = None
        for r in range(2, ws_best8.max_row + 1):
            name_val = ws_best8.cell(r,1).value
            grp = classify({"품목명": name_val or "", "중량": ws_best8.cell(r,2).value or ""})
            fill = drip_fill if grp == "드립백" else white_fill
            for c in ws_best8[r]:
                c.font = body_font; c.fill = fill; c.border = THIN_BORDER
        # 합계 행
        last = ws_best8.max_row + 1
        ws_best8.cell(last, 1, "합계").font = Font(name="Arial", size=10, bold=True)
        ws_best8.cell(last, 4, f"=SUM(D2:D{last-1})").font = Font(name="Arial", size=10, bold=True)
        for c in ws_best8[last]:
            c.fill = PatternFill("solid", fgColor=COLOR_HEADER); c.border = THIN_BORDER
    if not sheet3_df.empty:
        write_simple_sheet(ws3, sheet3_df, ["품목명", "빈칸", "이름", "수량"])
        insert_sheet3_into_sheet1(wb)
    postprocess_festa_rows(wb["주문취합"])
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ──────────────────────────────────────────────
# 분배모드 결과 → 엑셀 BytesIO
# ──────────────────────────────────────────────
def _parse_sku_app(sku):
    """배치 SKU 시트 및 품목준비 전송용 파싱 함수 (모듈 레벨)"""
    import re as _re
    WP = _re.compile(r'(\d+(?:\.\d+)?\s*(?:kg|g))', _re.IGNORECASE)
    RM = ['/필요','필요','/불필요','불필요','/구매 안함','구매 안함','/플러스','플러스']
    sku = str(sku).strip(); clean = sku
    for rm in RM:
        if clean.endswith(rm): clean = clean[:-len(rm)].strip().strip('/')
    if '드립백' in clean:
        if '_' in clean:
            n, r = clean.split('_', 1); r = r.strip().strip('/')
            for rm in RM: r = r.replace(rm, '').strip().strip('/')
            return n.strip(), '', r
        return clean, '', ''
    if '_' in clean:
        n, r = clean.split('_', 1); m = WP.search(r)
        if m:
            w = m.group(1).replace(' ', '')
            o = (r[:m.start()] + r[m.end():]).strip().strip('/')
        else: w = ''; o = r.strip()
    else: n = clean; w = ''; o = ''
    for rm in RM: o = o.replace(rm, '').strip().strip('/')
    return n.strip(), w, o

def build_das_excel(upload_df, line_df, seed_df, stats):
    import re as _re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    TERRA='8B3A2A'; LINEN='EDE5DC'; CREAM='FBF8F5'; OFF='F5F0EB'
    PEAK='F5E8D8'; CH='2C2C2C'; STONE='D6CEC8'
    thin=Side(style='thin',color=STONE); B=Border(left=thin,right=thin,top=thin,bottom=thin)
    WP = _re.compile(r'(\d+(?:\.\d+)?\s*(?:kg|g))', _re.IGNORECASE)
    RM = ['/구매 안함','구매 안함','/불필요','불필요','/플러스','플러스','/필요','필요']

    def parse_sku(sku):
        sku = str(sku).strip()
        # 불필요 문자 제거 (파싱 전)
        STRIP_SUFFIX = ['/필요', '필요', '/불필요', '불필요', '/구매 안함', '구매 안함',
                        '/플러스', '플러스']
        clean = sku
        for rm in STRIP_SUFFIX:
            if clean.endswith(rm):
                clean = clean[:-len(rm)].strip().strip('/')
        # 드립백: _ 뒤 전체가 옵션(B열), 중량 없음
        if '드립백' in clean:
            if '_' in clean:
                name, rest = clean.split('_', 1)
                rest = rest.strip().strip('/')
                # 맛 구성 뒤에 붙은 /필요 등 제거
                for rm in STRIP_SUFFIX:
                    rest = rest.replace(rm, '').strip().strip('/')
                return name.strip(), '', rest  # B열=공백, C열=옵션(맛구성)
            return clean, '', ''
        # 일반 상품: 중량 추출
        if '_' in clean:
            name, rest = clean.split('_', 1)
            m = WP.search(rest)
            if m:
                weight = m.group(1).replace(' ', '')
                option = (rest[:m.start()] + rest[m.end():]).strip().strip('/')
            else:
                weight = ''; option = rest.strip()
        else:
            name = clean; weight = ''; option = ''
        # 옵션 불필요 문자 제거
        for rm in STRIP_SUFFIX:
            option = option.replace(rm, '').strip().strip('/')
        return name.strip(), weight, option

    # 배치별 주문 건수 (칸배정 행에서 추출)
    slot_rows = seed_df[seed_df['SKU'].astype(str).str.startswith('[칸배정]')].copy()
    batch_order_counts = slot_rows.groupby('배치').size().to_dict()

    # 배치별 송장 순번 범위 추출
    def _get_serials(분배_str):
        import re as _re2
        m = _re2.search(r'송장순번\s*(\d+)', str(분배_str))
        return int(m.group(1)) if m else None

    batch_serial_ranges = {}
    for batch, grp in slot_rows.groupby('배치'):
        serials = grp['분배'].apply(_get_serials).dropna().astype(int)
        if not serials.empty:
            batch_serial_ranges[batch] = (int(serials.min()), int(serials.max()))

    # 배치별 SKU 데이터
    sku_rows = seed_df[~seed_df['SKU'].astype(str).str.startswith('[칸배정]')].copy()
    batch_sheets = {}
    batch_bags = {}
    for batch in sorted(sku_rows['배치'].unique()):
        rows = []
        for _, r in sku_rows[sku_rows['배치']==batch].iterrows():
            name, weight, option = parse_sku(r['SKU'])
            rows.append({'품목명': name, '중량': weight, '옵션': option, '수량': r['총수량']})
        df_b = pd.DataFrame(rows)
        # 동일 품목명+중량+옵션 합산
        df_b = df_b.groupby(['품목명','중량','옵션'], sort=False, as_index=False)['수량'].sum()

        # 쇼핑백 필요 수량 집계 (필요 키워드가 원본 SKU에 있던 행)
        bag_s = sum(r['총수량'] for _, r in sku_rows[sku_rows['배치']==batch].iterrows()
                    if '드립백' in str(r['SKU']) and '필요' in str(r['SKU']) and '10개입' in str(r['SKU']))
        bag_l = sum(r['총수량'] for _, r in sku_rows[sku_rows['배치']==batch].iterrows()
                    if '드립백' in str(r['SKU']) and '필요' in str(r['SKU']) and '30개입' in str(r['SKU']))
        bag_rows = []
        if bag_s: bag_rows.append({'품목명':'쇼핑백(소) 필요','중량':'','옵션':'','수량': bag_s})
        if bag_l: bag_rows.append({'품목명':'쇼핑백(대) 필요','중량':'','옵션':'','수량': bag_l})
        batch_bags[batch] = bag_rows
        # 주문취합과 동일한 정렬: 세트→기타→드립백→스쿱세트→원두, 품목명→중량 내림→옵션
        df_b['_group'] = df_b.apply(lambda r: classify({'품목명': r['품목명'], '중량': r['중량']}), axis=1)
        df_b['_g_order'] = df_b['_group'].map(GROUP_ORDER)
        df_b['_w_gram'] = df_b['중량'].apply(weight_to_gram)
        df_b = df_b.sort_values(['_g_order','품목명','_w_gram','옵션'],
                                ascending=[True,True,False,True]).drop(columns=['_group','_g_order','_w_gram'])
        batch_sheets[batch] = df_b.reset_index(drop=True)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        line_df.to_excel(w, sheet_name='줄포장(단일주문)', index=False)
        seed_df.to_excel(w, sheet_name='씨딩지시(복합주문)', index=False)
        for batch, df_b in batch_sheets.items():
            df_b.to_excel(w, sheet_name=f'배치{int(batch)}_SKU', index=False)
        pd.DataFrame([stats]).to_excel(w, sheet_name='요약', index=False)
    buf.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(buf)

    # 줄포장·씨딩지시 스타일
    for sn, widths in [('줄포장(단일주문)',[10,60,8,18]), ('씨딩지시(복합주문)',[6,60,8,70])]:
        ws = wb[sn]; ws.sheet_view.showGridLines=False; ws.freeze_panes='A2'
        for i,wd in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=wd
        for c in ws[1]:
            c.font=Font(name='맑은 고딕',bold=True,size=9,color='FFFFFF')
            c.fill=PatternFill('solid',start_color=TERRA)
            c.alignment=Alignment(horizontal='center'); c.border=B
        prev=None; band=False
        for r in range(2, ws.max_row+1):
            gval = ws.cell(row=r,column=2).value if sn=='줄포장(단일주문)' else ws.cell(row=r,column=1).value
            if gval!=prev: band=not band; prev=gval
            slot = sn=='씨딩지시(복합주문)' and str(ws.cell(row=r,column=2).value or '').startswith('[칸배정]')
            for c in ws[r]:
                c.font=Font(name='맑은 고딕',size=9,color='6B6056' if slot else CH); c.border=B
                c.fill=PatternFill('solid',start_color=PEAK if slot else (CREAM if band else OFF))

    # 배치별 SKU 시트 스타일
    for batch in sorted(sku_rows['배치'].unique()):
        sname = f'배치{int(batch)}_SKU'
        if sname not in wb.sheetnames: continue
        ws_b = wb[sname]; ws_b.sheet_view.showGridLines=False
        ws_b.column_dimensions['A'].width=42; ws_b.column_dimensions['B'].width=10
        ws_b.column_dimensions['C'].width=30; ws_b.column_dimensions['D'].width=8

        # 쇼핑백 행 삽입 (헤더 다음에 넣을 행 수 계산)
        bags = batch_bags.get(batch, [])
        n_bag = len(bags)

        # 타이틀 행 삽입 (1행)
        ws_b.insert_rows(1)
        order_cnt = batch_order_counts.get(batch, '?')
        sr = batch_serial_ranges.get(batch, None)
        serial_str = f'송장순번 {sr[0]}~{sr[1]}' if sr else ''
        title_val = f'배치 {int(batch)}  |  복합주문 {order_cnt}건  |  {serial_str}'
        tc = ws_b.cell(row=1, column=1, value=title_val)
        tc.font=Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
        tc.fill=PatternFill('solid', start_color=TERRA)
        tc.alignment=Alignment(vertical='center', indent=1)
        ws_b.row_dimensions[1].height=22
        ws_b.merge_cells(f'A1:D1')

        # 헤더 행 (2행)
        for c in ws_b[2]:
            c.font=Font(name='맑은 고딕',bold=True,size=9,color='FFFFFF')
            c.fill=PatternFill('solid',start_color=TERRA); c.border=B
            c.alignment=Alignment(horizontal='center')

        # 쇼핑백 행 삽입 (3행~) — 일반 데이터 행과 동일 스타일
        if n_bag:
            ws_b.insert_rows(3, amount=n_bag)
            for bi, bag in enumerate(bags):
                br = 3 + bi
                fill = PatternFill('solid', start_color=CREAM if bi%2==0 else OFF)
                ws_b.cell(row=br, column=1, value=bag['품목명']).font = Font(name='맑은 고딕', size=10, color=CH)
                ws_b.cell(row=br, column=4, value=bag['수량']).font = Font(name='맑은 고딕', size=10, color=CH)
                for col in range(1, 5):
                    c = ws_b.cell(row=br, column=col)
                    c.fill = fill; c.border = B
                ws_b.cell(row=br, column=4).alignment = Alignment(horizontal='center')

        ws_b.freeze_panes = f'A{3 + n_bag}'
        prev_name=None; band=False
        for r in range(3 + n_bag, ws_b.max_row+1):
            cur=ws_b.cell(row=r,column=1).value
            if cur!=prev_name: band=not band; prev_name=cur
            for c in ws_b[r]:
                c.font=Font(name='맑은 고딕',size=10,color=CH); c.border=B
                c.fill=PatternFill('solid',start_color=CREAM if band else OFF)
            ws_b.cell(row=r,column=4).alignment=Alignment(horizontal='center')
        # 합계 행
        last = ws_b.max_row+1
        for col in range(1,5):
            c=ws_b.cell(row=last,column=col)
            c.fill=PatternFill('solid',start_color=LINEN); c.border=B
        ws_b.cell(row=last,column=1,value='합계').font=Font(name='맑은 고딕',bold=True,size=10,color=TERRA)
        ws_b.cell(row=last,column=1).fill=PatternFill('solid',start_color=LINEN)
        tot=ws_b.cell(row=last,column=4,value=f'=SUM(D2:D{last-1})')
        tot.font=Font(name='맑은 고딕',bold=True,size=10,color=TERRA)
        tot.alignment=Alignment(horizontal='center')

    out = BytesIO(); wb.save(out); out.seek(0)
    return out


# ══════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════
st.set_page_config(page_title="테라로사 주문관리", page_icon="☕", layout="wide")

st.markdown("""
<style>
[data-testid="stSidebar"] { background: #FAF3F0; }
[data-testid="stSidebar"] .block-container { padding-top: 1.5rem; }
h1 { color: #8B3A2A !important; }
h2, h3 { color: #8B3A2A !important; }
.stButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600;
}
.stButton > button:hover { background: #C4644A; color: white; }
.stDownloadButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600; width: 100%;
}
</style>
""", unsafe_allow_html=True)

if "set_config" not in st.session_state:
    st.session_state.set_config = load_set_config()
if "editing_set" not in st.session_state:
    st.session_state.editing_set = None

# ─── 사이드바 (세트 상품 관리 — 기존 그대로) ───
with st.sidebar:
    st.markdown("## ☕ 세트 상품 관리")
    st.caption("세트 상품을 구성 품목별로 분리합니다.")
    st.divider()
    with st.expander("➕ 새 세트 상품 추가", expanded=False):
        new_set_name = st.text_input("세트 상품명", placeholder="예: 간편커피&유리머그 세트", key="new_set_name")
        if st.button("추가", key="btn_add_set", use_container_width=True):
            name = new_set_name.strip()
            if name and name not in st.session_state.set_config:
                st.session_state.set_config[name] = []
                save_set_config(st.session_state.set_config)
                st.session_state.editing_set = name
                st.rerun()
            elif name in st.session_state.set_config:
                st.warning("이미 등록된 세트 상품입니다.")
            else:
                st.warning("세트 상품명을 입력하세요.")
    st.divider()
    if not st.session_state.set_config:
        st.info("등록된 세트 상품이 없습니다.")
    else:
        st.markdown(f"**등록된 세트** {len(st.session_state.set_config)}개")
        for set_name in list(st.session_state.set_config.keys()):
            comps = st.session_state.set_config[set_name]
            is_editing = st.session_state.editing_set == set_name
            with st.container():
                col_name, col_edit, col_del = st.columns([6, 2, 2])
                with col_name:
                    st.markdown(f"**{set_name}**")
                    st.caption(f"구성 {len(comps)}개")
                with col_edit:
                    if st.button("편집" if not is_editing else "닫기", key=f"edit_{set_name}", use_container_width=True):
                        st.session_state.editing_set = None if is_editing else set_name
                        st.rerun()
                with col_del:
                    if st.button("삭제", key=f"del_{set_name}", use_container_width=True, type="secondary"):
                        del st.session_state.set_config[set_name]
                        if st.session_state.editing_set == set_name:
                            st.session_state.editing_set = None
                        save_set_config(st.session_state.set_config)
                        st.rerun()
                if is_editing:
                    with st.container():
                        st.markdown(f"###### 구성 품목 — {set_name}")
                        for i, comp in enumerate(comps):
                            c1, c2, c3, c4, c5 = st.columns([1, 4, 2, 2, 1])
                            with c1:
                                new_qty = st.number_input("수량", min_value=1, value=comp.get("qty", 1), key=f"qty_{set_name}_{i}", label_visibility="collapsed")
                            with c2:
                                new_name = st.text_input("품목명", value=comp.get("name", ""), key=f"cname_{set_name}_{i}", label_visibility="collapsed")
                            with c3:
                                new_weight = st.text_input("중량", value=comp.get("weight", ""), placeholder="예: 250g", key=f"cweight_{set_name}_{i}", label_visibility="collapsed")
                            with c4:
                                new_option = st.text_input("옵션", value=comp.get("option", ""), placeholder="옵션(선택)", key=f"coption_{set_name}_{i}", label_visibility="collapsed")
                            with c5:
                                if st.button("✕", key=f"rm_{set_name}_{i}"):
                                    comps.pop(i)
                                    save_set_config(st.session_state.set_config)
                                    st.rerun()
                            comps[i] = {"name": new_name, "qty": int(new_qty), "weight": new_weight, "option": new_option}
                        st.markdown("---")
                        na1, na2, na3, na4, na5 = st.columns([1, 4, 2, 2, 1])
                        with na1:
                            add_qty = st.number_input("수량", min_value=1, value=1, key=f"addqty_{set_name}", label_visibility="collapsed")
                        with na2:
                            add_name = st.text_input("품목명", placeholder="구성 품목명", key=f"addname_{set_name}", label_visibility="collapsed")
                        with na3:
                            add_weight = st.text_input("중량", placeholder="예: 250g", key=f"addweight_{set_name}", label_visibility="collapsed")
                        with na4:
                            add_option = st.text_input("옵션", placeholder="옵션(선택)", key=f"addoption_{set_name}", label_visibility="collapsed")
                        with na5:
                            if st.button("＋", key=f"addcomp_{set_name}"):
                                if add_name.strip():
                                    comps.append({"name": add_name.strip(), "qty": int(add_qty), "weight": add_weight.strip(), "option": add_option.strip()})
                                    save_set_config(st.session_state.set_config)
                                    st.rerun()
                        if st.button("💾 저장", key=f"save_{set_name}", use_container_width=True):
                            save_set_config(st.session_state.set_config)
                            st.success("저장 완료!")
                st.divider()

# ─── 메인: 탭 2개 ───
st.title("테라로사 주문관리")
tab1, tab2 = st.tabs(["📋 주문취합", "📦 분배모드 (DAS)"])

# ════ 탭1: 주문취합 (기존 그대로) ════
with tab1:
    if st.session_state.set_config:
        with st.expander(f"📦 세트 분리 설정 — {len(st.session_state.set_config)}개 등록됨", expanded=False):
            cols = st.columns(min(3, len(st.session_state.set_config)))
            for i, (sname, comps) in enumerate(st.session_state.set_config.items()):
                with cols[i % 3]:
                    st.markdown(f"**{sname}**")
                    for c in comps:
                        weight_str = f" {c['weight']}" if c.get("weight") else ""
                        option_str = f" / {c['option']}" if c.get("option") else ""
                        st.caption(f"× {c['qty']}  {c['name']}{weight_str}{option_str}")
    st.divider()
    col1, col2 = st.columns(2)
    with col1:
        order_file = st.file_uploader("📄 주문취합 Excel", type=["xlsx"], help="'취합용' 시트가 포함된 주문 파일", key="order_file")
    with col2:
        code_file = st.file_uploader("📋 자사몰 상품코드 Excel", type=["xlsx"], help="상품코드 매핑 파일", key="code_file")
    st.divider()
    if order_file and code_file:
        if st.button("🚀 주문 취합 처리 시작", use_container_width=True, key="btn_process"):
            with st.spinner("처리 중..."):
                try:
                    result_buf = process(order_file, code_file, st.session_state.set_config)
                    today = datetime.today().strftime("%Y%m%d")
                    st.success("✅ 처리 완료!")
                    set_count = len(st.session_state.set_config)
                    if set_count:
                        st.info(f"📦 세트 분리 적용: {set_count}개 세트 상품 → 구성 품목별 행으로 분리됨")
                    st.download_button(
                        label="⬇️ 결과 Excel 다운로드",
                        data=result_buf,
                        file_name=f"자사몰주문취합_{today}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_order"
                    )
                except Exception as e:
                    st.error(f"❌ 오류 발생: {e}")
                    st.exception(e)
    else:
        st.info("👆 주문취합 파일과 상품코드 파일을 모두 업로드하면 처리 버튼이 활성화됩니다.")

# ════ 탭2: 분배모드 (DAS) ════
with tab2:
    st.markdown("### 📦 분배모드 (DAS)")
    st.caption("택배용(발송양식) 파일을 업로드하면 우체국 업로드용 정렬본과 현장 작업지시서를 만들어줍니다. 개인정보는 이 PC 밖으로 전송되지 않습니다.")
    st.divider()

    col_a, col_b = st.columns([3, 1])
    with col_a:
        das_file = st.file_uploader(
            "📄 택배용 발송양식 파일 (xlsx)",
            type=["xlsx"],
            help="사방넷에서 다운로드한 택배용 시트가 있는 파일",
            key="das_file"
        )
    with col_b:
        batch_size = st.number_input("선반 칸수", min_value=8, max_value=200, value=40, step=1,
                                      help="DAS 선반 슬롯 수 (기본 40)")

    # 결과 session_state 초기화
    if "das_result" not in st.session_state:
        st.session_state.das_result = None

    if das_file:
        # 파일 검증
        try:
            xl = pd.ExcelFile(das_file)
            sheet = '택배용' if '택배용' in xl.sheet_names else xl.sheet_names[0]
            df_das = pd.read_excel(das_file, sheet_name=sheet).dropna(how='all')
            ITEM_COL = '상품명(확정)+옵션(확정)+수량(조합용)'
            if ITEM_COL not in df_das.columns:
                st.error("❌ 택배용(발송양식) 파일이 아닙니다. 올바른 파일을 업로드하세요.")
                st.stop()
            st.success(f"✅ 파일 인식 완료 — '{sheet}' 시트, {len(df_das)}행")
        except Exception as e:
            st.error(f"❌ 파일 읽기 오류: {e}")
            st.stop()

        if st.button("🚀 분배모드 처리 시작", use_container_width=True, key="btn_das"):
            with st.spinner("처리 중..."):
                try:
                    upload_df, line_df, seed_df, stats, _ = das_run(df_das, batch_size=int(batch_size))
                    today = datetime.today().strftime("%m%d_%H%M")
                    das_buf = build_das_excel(upload_df, line_df, seed_df, stats)
                    # 전화번호 열 텍스트 형식으로 저장 (앞자리 0 보존)
                    from openpyxl import load_workbook as _lw
                    up_buf = BytesIO()
                    upload_df.to_excel(up_buf, index=False)
                    up_buf.seek(0)
                    _wb = _lw(up_buf)
                    _ws = _wb.active
                    # 헤더에서 전화번호 열 찾기
                    _phone_cols = []
                    for _ci, _cell in enumerate(_ws[1], 1):
                        if _cell.value and '전화번호' in str(_cell.value):
                            _phone_cols.append(_ci)
                    # 해당 열 텍스트 서식 + 앞자리 0 패딩
                    for _row in _ws.iter_rows(min_row=2):
                        for _ci in _phone_cols:
                            _c = _row[_ci - 1]
                            if _c.value is not None:
                                _v = str(int(_c.value)) if isinstance(_c.value, float) else str(_c.value)
                                # 11자리 미만이면 앞에 0 패딩
                                if _v.isdigit() and len(_v) < 11:
                                    _v = _v.zfill(11)
                                _c.value = _v
                                _c.number_format = '@'
                    up_buf = BytesIO()
                    _wb.save(up_buf)
                    up_buf.seek(0)

                    # 씨딩 데이터 추출 (개인정보 없음 — SKU/칸번호/수량만)
                    seed_sku_rows = seed_df[~seed_df['SKU'].astype(str).str.startswith('[칸배정]')].to_dict('records')

                    # 결과를 session_state에 저장
                    st.session_state.das_result = {
                        "stats": stats,
                        "up_buf": up_buf.getvalue(),
                        "das_buf": das_buf.getvalue(),
                        "seed_df": seed_df,
                        "seed_sku_rows": seed_sku_rows,
                        "today": today,
                    }
                except Exception as e:
                    st.error(f"❌ 오류 발생: {e}")
                    st.exception(e)

        # 결과가 있으면 항상 표시 (다운로드 버튼 눌러도 사라지지 않음)
        if st.session_state.das_result:
            r = st.session_state.das_result
            stats = r["stats"]
            st.success("✅ 처리 완료! 검증 3종 통과")
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("총 주문", stats['총주문'])
            c2.metric("줄포장", stats['단일'])
            c3.metric("씨딩(복합)", stats['복합'])
            c4.metric("배치 수", stats['배치수'])
            c5.metric("SKU 방문", stats['씨딩SKU방문'])
            st.divider()
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="⬇️ ①업로드용 정렬본 다운로드",
                    data=r["up_buf"],
                    file_name=f"①업로드용_정렬본_{r['today']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_upload"
                )
                st.caption("우체국 사이트에 그대로 업로드")
            with col2:
                st.download_button(
                    label="⬇️ ②작업지시서 다운로드",
                    data=r["das_buf"],
                    file_name=f"②작업지시서_{r['today']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_guide"
                )
                st.caption("줄포장 리스트 / 씨딩지시 / 요약")
            # 전체 씨딩 데이터 한 번에 전송 (로컬 전용 — 개인정보 없는 씨딩 데이터만 GitHub에 전송)
            if "seed_sku_rows" in r:
                st.divider()
                st.markdown("**📱 태블릿 씨딩 화면으로 전송**")
                st.caption("SKU명·칸번호·수량만 전송됩니다. 개인정보는 포함되지 않습니다.")
                # 전체 씨딩 전송
                if st.button("📦 태블릿으로 전송 (씨딩 + 품목 준비)", use_container_width=True, key="btn_send_all"):
                    seed_df_r = r["seed_df"]
                    batches   = sorted(seed_df_r['배치'].unique())

                    # 씨딩 데이터
                    all_batches = {}
                    for b in batches:
                        b_skus = [row for row in r["seed_sku_rows"] if row['배치'] == b]
                        all_batches[str(int(b))] = b_skus

                    # 배치 시트 데이터 (작업지시서 배치N_SKU 시트와 동일)
                    all_prep = {}
                    for b in batches:
                        b_rows = []
                        for _, row in seed_df_r[
                            (seed_df_r['배치']==b) &
                            (~seed_df_r['SKU'].astype(str).str.startswith('[칸배정]'))
                        ].iterrows():
                            name, weight, option = _parse_sku_app(str(row['SKU']))
                            b_rows.append({'품목명': name, '중량': weight, '옵션': option, '수량': int(row['총수량'])})
                        mask_base = (
                            (seed_df_r['배치']==b) &
                            (~seed_df_r['SKU'].astype(str).str.startswith('[칸배정]')) &
                            seed_df_r['SKU'].str.contains('드립백', na=False) &
                            seed_df_r['SKU'].str.contains('필요', na=False)
                        )
                        bag_s = int(seed_df_r[mask_base & seed_df_r['SKU'].str.contains('10개입', na=False)]['총수량'].sum())
                        bag_l = int(seed_df_r[mask_base & seed_df_r['SKU'].str.contains('30개입', na=False)]['총수량'].sum())
                        df_b  = pd.DataFrame(b_rows)
                        df_b  = df_b.groupby(['품목명','중량','옵션'], sort=False, as_index=False)['수량'].sum()
                        df_b['_group']   = df_b.apply(lambda r: classify({'품목명': r['품목명'], '중량': r['중량']}), axis=1)
                        df_b['_g_order'] = df_b['_group'].map(GROUP_ORDER)
                        df_b['_w_gram']  = df_b['중량'].apply(weight_to_gram)
                        df_b = df_b.sort_values(['_g_order','품목명','_w_gram','옵션'],
                                                ascending=[True,True,False,True]).drop(columns=['_group','_g_order','_w_gram'])
                        b_rows = df_b.to_dict('records')
                        if bag_l: b_rows.insert(0, {'품목명':'쇼핑백(대) 필요','중량':'','옵션':'','수량': bag_l})
                        if bag_s: b_rows.insert(0, {'품목명':'쇼핑백(소) 필요','중량':'','옵션':'','수량': bag_s})
                        # 배치 메타 정보 추가 (타이틀용)
                        slot_rows_b = seed_df_r[(seed_df_r['배치']==b) & seed_df_r['SKU'].astype(str).str.startswith('[칸배정]')]
                        order_cnt   = len(slot_rows_b)
                        import re as _re
                        serials = slot_rows_b['분배'].apply(lambda x: int(m.group(1)) if (m:=_re.search(r'송장순번\s*(\d+)', str(x))) else None).dropna().astype(int)
                        serial_range = f'{serials.min()}~{serials.max()}' if not serials.empty else ''
                        all_prep[str(int(b))] = {
                            'items': b_rows,
                            'order_cnt': order_cnt,
                            'serial_range': serial_range,
                            'total_qty': sum(r['수량'] for r in b_rows),
                        }

                    # 하나의 파일로 통합 전송
                    ok = gh_save("das_data.json", {
                        "batch_list": [int(b) for b in batches],
                        "batches":    all_batches,
                        "all_prep":   all_prep,
                    })
                    if ok:
                        st.success(f"✅ {len(batches)}개 배치 전송 완료! 태블릿에서 '불러오기'를 누르세요.")
                    else:
                        st.error("❌ 전송 실패. GitHub 연결을 확인하세요.")

            with st.expander("씨딩지시 미리보기", expanded=False):
                st.dataframe(r["seed_df"], use_container_width=True, height=400)
    else:
        st.session_state.das_result = None
        st.info("👆 택배용 발송양식 파일을 업로드하면 처리 버튼이 활성화됩니다.")

    with st.expander("ℹ️ 분배모드 사용법"):
        st.markdown("""
**분배모드란?**
주문을 단일 SKU / 복합(2종 이상)으로 분류해 줄포장과 DAS 씨딩 작업을 최적화합니다.

**작업 순서**
1. 사방넷에서 택배용 발송양식 파일 다운로드
2. 이 화면에 업로드 → 처리 시작
3. **①업로드용 정렬본** → 우체국 사이트에 그대로 접수 (행 순서 = 송장 순서)
4. **②작업지시서** → 줄포장 리스트 순서대로 포장 → 씨딩지시 순서대로 복합주문 분배

**선반 칸수**
- 현재 보유 슬롯 수에 맞게 조정 (기본 40)
- 숫자가 클수록 SKU 방문 횟수가 줄어 씨딩 효율 증가
        """)
